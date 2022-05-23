import requests
from setuptools import setup
from classes.database_access import DB_Connect
import csv
import openpyxl
import ezsheets
import time
import threading

#Connects to the database being used
my_db = DB_Connect('root', '', 'scheduling_tasks_assignment')

#Imports the data from a webpage and writes it to a csv file stored in text_files
request = requests.get('https://ool-content.walshcollege.edu/CourseFiles/IT/IT414/MASTER/Week08/WI20-Assignment/exam_data.csv')

if request.status_code == 200:
    exam_data = open('text_files/exam_data.csv', 'wb')

    for chunk in request.iter_content(100000):
        exam_data.write(chunk)

    with open('text_files/script_log.txt', 'a') as log_file:
        log_file.write(time.ctime() + " - exam_data.csv retrieval complete.\n")

#Writes the data from the file to the database
def writeToDatabase():
    try:
        with open("text_files/exam_data.csv") as csv_file:
            #Clears existing data in working table
            my_db.executeQuery("TRUNCATE exam_data")

            #Reads data from csv file and adds it to a list
            csvReader = csv.reader(csv_file)
            csv_data = list(csvReader)
            
            #Fills working table with imported csv data. Skips first row containing headers.
            row_count = 1
            while row_count < len(csv_data):
                my_db.executeQuery("INSERT INTO exam_data (parental_education_level, test_prep_course, math_score, reading_score, writing_score) VALUES (\"" + csv_data[row_count][0] + "\",'" + csv_data[row_count][1] + "','" + csv_data[row_count][2] + "','" + csv_data[row_count][3] + "','" + csv_data[row_count][4] + "')")
                row_count += 1
            
            my_db.conn.commit()

            with open('text_files/script_log.txt', 'a') as log_file:
                log_file.write(time.ctime() + " - Database writing complete.\n")
    except:
        print("File not found.")
        exit()

#Creates an excel spreadsheet and writes the csv data to it
def createExcelSheet():
    #Creates an excel spreadsheet
    exam_workbook = openpyxl.Workbook()
    curr_sheet = exam_workbook.active
    curr_sheet.title = "Exam Data"

    #Adds data to spreadsheet
    row_count = 1

    #Adds headers
    with open("text_files/exam_data.csv") as csv_file:
        #Reads data from csv file and adds it to a list
        csvReader = csv.reader(csv_file)
        csv_data = list(csvReader)

        curr_sheet.cell(row=row_count, column=1).value = csv_data[row_count-1][0]
        curr_sheet.cell(row=row_count, column=2).value = csv_data[row_count-1][1]
        curr_sheet.cell(row=row_count, column=3).value = csv_data[row_count-1][2]
        curr_sheet.cell(row=row_count, column=4).value = csv_data[row_count-1][3]
        curr_sheet.cell(row=row_count, column=5).value = csv_data[row_count-1][4]

        row_count += 1

        while row_count <= len(csv_data):
            curr_sheet.cell(row=row_count, column=1).value = csv_data[row_count-1][0]
            curr_sheet.cell(row=row_count, column=2).value = csv_data[row_count-1][1]
            curr_sheet.cell(row=row_count, column=3).value = int(csv_data[row_count-1][2])
            curr_sheet.cell(row=row_count, column=4).value = int(csv_data[row_count-1][3])
            curr_sheet.cell(row=row_count, column=5).value = int(csv_data[row_count-1][4])

            row_count += 1

    #Changes column formatting
    curr_sheet.column_dimensions['A'].width = 25
    curr_sheet.column_dimensions['B'].width = 22
    curr_sheet.column_dimensions['C'].width = 12.5
    curr_sheet.column_dimensions['D'].width = 12.5
    curr_sheet.column_dimensions['E'].width = 12.5

    #Saves excel spreadsheet to text_files
    exam_workbook.save("text_files/exam_data.xlsx")
    with open('text_files/script_log.txt', 'a') as log_file:
        log_file.write(time.ctime() + " - Excel spreadsheet creation complete.\n")

#Creates a google spreadsheet and writes the csv data to it
def createGoogleSheet():
    #Creates google spreadsheet
    google_sheet = ezsheets.createSpreadsheet("Exam Data")
    curr_sheet = google_sheet[0]

    #Moves data into a list for each column
    with open("text_files/exam_data.csv") as csv_file:
        #Reads data from csv file and adds it to a list
        csvReader = csv.reader(csv_file)
        csv_data = list(csvReader)

        parent_education = []
        test_prep = []
        math_score = []
        reading_score = []
        writing_score = []

        row_count = 0

        while row_count < len(csv_data):
            parent_education.append(csv_data[row_count][0])
            test_prep.append(csv_data[row_count][1])
            math_score.append(csv_data[row_count][2])
            reading_score.append(csv_data[row_count][3])
            writing_score.append(csv_data[row_count][4])

            row_count += 1

    #Updates each column in the google sheet with the column lists
    curr_sheet.updateColumn(1, parent_education)
    curr_sheet.updateColumn(2, test_prep)
    curr_sheet.updateColumn(3, math_score)
    curr_sheet.updateColumn(4, reading_score)
    curr_sheet.updateColumn(5, writing_score)


    with open('text_files/script_log.txt', 'a') as log_file:
        log_file.write(time.ctime() + " - Google spreadsheet creation complete.\n")

#Utilizes multithreading to perform three functions
dbThread = threading.Thread(target=writeToDatabase)
excelThread = threading.Thread(target=createExcelSheet)
googleThread = threading.Thread(target=createGoogleSheet)

dbThread.start()
excelThread.start()
googleThread.start()
