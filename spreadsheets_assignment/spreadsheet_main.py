from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import *
import ezsheets

pencil_total = 0
binder_total = 0
pen_total = 0
desk_total = 0
pen_set_total = 0

#Creates a spreadsheet
my_workbook = openpyxl.Workbook()
curr_sheet = my_workbook.active
curr_sheet.title = "Raw Data"

#Obtains data from webpage
request = requests.get("https://ool-content.walshcollege.edu/CourseFiles/IT/IT414/MASTER/Week05/WI20-Assignment/sales_data.html")

if request.status_code == 200:
    sheet_data = open("files/sheet_data.html", "wb")
    for data_chunk in request.iter_content(100000):
        sheet_data.write(data_chunk)

sales_sheet = open("files/sheet_data.html")
sheet_soup = BeautifulSoup(sales_sheet, "lxml")

sheet_cells = sheet_soup.findAll("td")

#Adds data to spreadsheet
col_counter = 0
row_counter = 1

for item in sheet_cells:
    col_counter += 1
    if col_counter == 4:
        curr_sheet.cell(row=row_counter, column=1).value = item.getText()
    if col_counter == 5:
        curr_sheet.cell(row=row_counter, column=2).value = item.getText()
        if curr_sheet.cell(row=row_counter,column=1).value.strip() == "Pencil":
            pencil_total += int(item.getText())
        elif curr_sheet.cell(row=row_counter,column=1).value.strip() == "Binder":
            binder_total += int(item.getText())
        elif curr_sheet.cell(row=row_counter,column=1).value.strip() == "Pen":
            pen_total += int(item.getText())
        elif curr_sheet.cell(row=row_counter,column=1).value.strip() == "Desk":
            desk_total += int(item.getText())
        elif curr_sheet.cell(row=row_counter,column=1).value.strip() == "Pen Set":
            pen_set_total += int(item.getText())
    elif col_counter == 7:
        col_counter = 0
        row_counter += 1    

#Adds another sheet for calculated data
my_workbook.create_sheet("Total Quantity Sold")
curr_sheet = my_workbook["Total Quantity Sold"]

#Adds spreadsheet title and formatting
curr_sheet.merge_cells("A1:C1")
curr_sheet["A1"].value = "Sales Data"
curr_sheet["A1"].font = Font(sz=20.0, b=True, color="33E8FF")
curr_sheet["A1"].alignment = Alignment(wrap_text=True, horizontal="center")
curr_sheet.row_dimensions[1].height = 25

#Adds column headers and formatting
curr_sheet["A2"].value = "Product"
curr_sheet["A2"].font = Font(b=True)
curr_sheet["B2"].value = "Quantity"
curr_sheet["B2"].font = Font(b=True)

#Adds data to the sheet
curr_sheet["A3"].value = "Pencil"
curr_sheet["B3"].value = pencil_total
curr_sheet["A4"].value = "Binder"
curr_sheet["B4"].value = binder_total
curr_sheet["A5"].value = "Pen"
curr_sheet["B5"].value = pen_total
curr_sheet["A6"].value = "Desk"
curr_sheet["B6"].value = desk_total
curr_sheet["A7"].value = "Pen Set"
curr_sheet["B7"].value = pen_set_total

curr_sheet["A9"].value = "Avg:"
curr_sheet["B9"].value = "=AVERAGE(B3:B7)"

#Saves sheet in files folder
my_workbook.save("files/my_workbook.xlsx")

#Creates google spreadsheet and adds same data to it
google_sheet = ezsheets.createSpreadsheet("Product Sales")

curr_sheet = google_sheet[0]
rows = curr_sheet.getRows()

rows[0][0] = "Product"
rows[0][1] = "Quantity"

rows[1][0] = "Pencil"
rows[1][1] = pencil_total

rows[2][0] = "Binder"
rows[2][1] = binder_total

rows[3][0] = "Pen"
rows[3][1] = pen_total

rows[4][0] = "Desk"
rows[4][1] = desk_total

rows[5][0] = "Pen Set"
rows[5][1] = pen_set_total

rows[6][0] = "Avg:"
rows[6][1] = "=AVERAGE(B2:B6)"

curr_sheet.updateRows(rows)
