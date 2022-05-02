from distutils.cygwinccompiler import Mingw32CCompiler
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import *
from openpyxl.chart import Reference, PieChart

gradebook = openpyxl.load_workbook("files/gradebook.xlsx")
gradebook_list = []

curr_sheet = gradebook["Grades"]

max_row = curr_sheet.max_row
max_col = curr_sheet.max_column

#indices for rows/columns in spreadsheets start with 1 rather than 0
row_count = 1

while row_count <= max_row:
    col_count = 1
    tmp_row = {}

    while col_count <= max_col:
        tmp_row["col" + str(col_count)] = curr_sheet.cell(row=row_count, column=col_count).value
        col_count += 1

    gradebook_list.append(tmp_row)
    row_count += 1

gradebook.remove(gradebook["Grades"])

gradebook.create_sheet("Grades")
curr_sheet = gradebook["Grades"]

curr_sheet.merge_cells("A1:K1")
curr_sheet["A1"] = "Course Grading Data"

curr_sheet["A1"].font = Font(sz=20.0, b=True, color="33E8FF")
curr_sheet["A1"].alignment = Alignment(horizontal="center")

row_count = 3

for item_row in gradebook_list:
    col_count = 1
    for key, value in item_row.items():
        curr_sheet.cell(row=row_count, column=col_count).value = value

        if row_count == 3:
            curr_sheet.cell(row=row_count, column=col_count).font = Font(b=True)
            curr_sheet.cell(row=row_count, column=col_count).alignment = Alignment(wrap_text=True, horizontal="center")
            col_letter = get_column_letter(col_count)
            curr_sheet.column_dimensions[col_letter].width = 15
        
        col_count += 1
    row_count += 1

#Adding new columns
last_col_letter = get_column_letter(max_col)
total_col_letter = get_column_letter(max_col + 1)
percent_col_letter = get_column_letter(max_col + 2)
grade_col_letter = get_column_letter(max_col + 3)

curr_sheet.cell(row=3, column=max_col + 1).font = Font(b=True)
curr_sheet.cell(row=3, column=max_col + 1).alignment = Alignment(wrap_text=True, horizontal="center")
curr_sheet.column_dimensions[total_col_letter].width = 15
curr_sheet.cell(row=3, column=max_col + 1).value = "Total"

curr_sheet.cell(row=3, column=max_col + 2).font = Font(b=True)
curr_sheet.cell(row=3, column=max_col + 2).alignment = Alignment(wrap_text=True, horizontal="center")
curr_sheet.column_dimensions[percent_col_letter].width = 15
curr_sheet.cell(row=3, column=max_col + 2).value = "Percent"

curr_sheet.cell(row=3, column=max_col + 3).font = Font(b=True)
curr_sheet.cell(row=3, column=max_col + 3).alignment = Alignment(wrap_text=True, horizontal="center")
curr_sheet.column_dimensions[grade_col_letter].width = 15
curr_sheet.cell(row=3, column=max_col + 3).value = "Letter Grade"

#Adding data to columns
row_count = 4

while row_count <= max_row + 2:
    #Changes value of total column rows to "=SUM(C4:H4)"
    curr_sheet[total_col_letter + str(row_count)].value = "=SUM(C" + str(row_count) + ":" + last_col_letter + str(row_count) + ")"
    curr_sheet[percent_col_letter + str(row_count)].value = "=ROUND(((" + total_col_letter + str(row_count) + "/325) * 100), 2)"
    curr_sheet[grade_col_letter + str(row_count)].value = "=IF(" + percent_col_letter + str(row_count) + ">=90,\"A\",IF(" + percent_col_letter + str(row_count) + ">=80,\"B\",IF(" + percent_col_letter + str(row_count) + ">=70,\"C\",IF(" + percent_col_letter + str(row_count) + ">=60,\"D\",\"F\"))))"
   
    row_count += 1

#Adds a new row
new_max_row = curr_sheet.max_row + 1
curr_sheet.merge_cells("A" + str(new_max_row) + ":B" + str(new_max_row))
curr_sheet["A" + str(new_max_row)].font = Font(b=True)
curr_sheet["A" + str(new_max_row)].alignment = Alignment(wrap_text=True, horizontal="right")
curr_sheet["A" + str(new_max_row)].value = "Average:"

avg_col_start = 3

while avg_col_start <= (max_col + 2):
    tmp_col = get_column_letter(avg_col_start)
    curr_sheet.cell(row=new_max_row, column=avg_col_start).value = "=ROUND((AVERAGE(" + tmp_col + str(4) + ":" + tmp_col + str(new_max_row - 1) + ")), 2)"
    avg_col_start += 1

curr_sheet.freeze_panes = "A4"

#Making a pie chart on a new sheet
gradebook.create_sheet("Graph")
curr_sheet = gradebook["Graph"]

curr_sheet["A1"] = "Grade"
curr_sheet["B1"] = "Frequency"
curr_sheet["A2"] = "A"
curr_sheet["B2"] = "=COUNTIF(Grades!" + grade_col_letter + str(4) + ":" + grade_col_letter + str(new_max_row - 1) + ", A2)"
curr_sheet["A3"] = "B"
curr_sheet["B3"] = "=COUNTIF(Grades!" + grade_col_letter + str(4) + ":" + grade_col_letter + str(new_max_row - 1) + ", A3)"
curr_sheet["A4"] = "C"
curr_sheet["B4"] = "=COUNTIF(Grades!" + grade_col_letter + str(4) + ":" + grade_col_letter + str(new_max_row - 1) + ", A4)"
curr_sheet["A5"] = "D"
curr_sheet["B5"] = "=COUNTIF(Grades!" + grade_col_letter + str(4) + ":" + grade_col_letter + str(new_max_row - 1) + ", A5)"
curr_sheet["A6"] = "F"
curr_sheet["B6"] = "=COUNTIF(Grades!" + grade_col_letter + str(4) + ":" + grade_col_letter + str(new_max_row - 1) + ", A6)"

pie_chart = PieChart()
chart_labels = Reference(curr_sheet, min_col=1, min_row=2, max_row=6)
chart_data = Reference(curr_sheet, min_col=2, min_row=1, max_row=6)
pie_chart.add_data(chart_data, titles_from_data=True)
pie_chart.set_categories(chart_labels)
pie_chart.title = "Grade Distributions"

curr_sheet.add_chart(pie_chart, "A9")

gradebook.save("files/gradebook_reformat.xlsx")