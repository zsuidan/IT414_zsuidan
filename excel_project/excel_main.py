from tokenize import maybe
import openpyxl

my_workbook = openpyxl.Workbook()
curr_sheet = my_workbook.active
curr_sheet.title = "My Sheet"
my_workbook.create_sheet(title = "Another Sheet")
curr_sheet = my_workbook["Another Sheet"]
curr_sheet["B3"] = "My Cell"

birthdays = [{"name": "Chris", "birth_date": "1/23/1975"},
            {"name": "Diane", "birth_date": "2/23/1982"},
            {"name": "Jamal", "birth_date": "4/21/1980"},
            {"name": "Padma", "birth_date": "11/07/1987"}]

#indices for rows/columns in spreadsheets start with 1 rather than 0
curr_sheet = my_workbook["My Sheet"]
row_count = 1

#Loops through each dictionary in the birthdays list
for item in birthdays:
    #Sets column to the first column
    col_count = 1
    #Adds name info to first column
    curr_sheet.cell(row = row_count, column = col_count).value = item["name"]
    #Changes to second column
    col_count += 1
    #Adds birth date info to secodn column
    curr_sheet.cell(row = row_count, column = col_count).value = item["birth_date"]
    #Moves to next row
    row_count += 1

#Changes width of B column to 50
curr_sheet.column_dimensions["B"].width = 50

#Saves the workbook as an exceel sheet in the files folder
my_workbook.save("files/my_workbook.xlsx")